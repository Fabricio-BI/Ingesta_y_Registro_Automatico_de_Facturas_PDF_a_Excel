"""
Exportación de resultados a un archivo Excel.
"""
import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from plantillas.esquema import CAMPOS_FACTURA

logger = logging.getLogger(__name__)

COLUMNAS = CAMPOS_FACTURA + ["plantilla_usada", "estado", "errores"]

FUENTE_TITULO = Font(name="Arial", bold=True, color="FFFFFF")
RELLENO_TITULO = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FUENTE_NORMAL = Font(name="Arial")
RELLENO_ERROR = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")


def exportar_a_excel(resultados: list[dict], ruta_salida: str) -> None:
    """
    Escribe una lista de resultados de facturas a un archivo .xlsx.

    Args:
        resultados: lista de diccionarios, cada uno con los campos de
            CAMPOS_FACTURA más 'plantilla_usada', 'estado' y 'errores'.
        ruta_salida: ruta donde se guardará el archivo Excel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"

    # Encabezados
    for col_idx, nombre_columna in enumerate(COLUMNAS, start=1):
        celda = ws.cell(row=1, column=col_idx, value=nombre_columna.replace("_", " ").title())
        celda.font = FUENTE_TITULO
        celda.fill = RELLENO_TITULO

    # Filas de datos
    for fila_idx, resultado in enumerate(resultados, start=2):
        es_valida = resultado.get("estado") == "válida"
        for col_idx, nombre_columna in enumerate(COLUMNAS, start=1):
            valor = resultado.get(nombre_columna, "")
            celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
            celda.font = FUENTE_NORMAL
            if not es_valida:
                celda.fill = RELLENO_ERROR

    # Ancho de columnas legible
    for col_idx, nombre_columna in enumerate(COLUMNAS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(nombre_columna) + 4)

    try:
        wb.save(ruta_salida)
    except PermissionError:
        base, ext = os.path.splitext(ruta_salida)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta_alternativa = f"{base}_{timestamp}{ext}"
        
        logger.error(
            f"¡ALERTA!: No se pudo guardar el archivo en '{ruta_salida}' porque está abierto por otro programa o usuario.\n"
            f"Se ha guardado una copia segura en: '{ruta_alternativa}'"
        )
        wb.save(ruta_alternativa)
